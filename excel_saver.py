import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
import logger_util


def _norm_phone(s: str) -> str:
    """Zwraca numer złożony wyłącznie z cyfr (do porównań)."""
    return ''.join(ch for ch in str(s) if ch.isdigit())


def save_to_excel(new_data, filename="firmy.xlsx"):
    today = datetime.today().strftime('%Y-%m-%d')

    if os.path.exists(filename):
        logger_util.log_info(f"✅ Otwieramy {filename}")
        book = load_workbook(filename)
    else:
        logger_util.log_info(f"✅ Tworzymy nowy {filename}")
        book = Workbook()

    # Pobierz lub utwórz arkusz dzisiejszy
    if today in book.sheetnames:
        sheet = book[today]
        try:
            existing_df = pd.read_excel(filename, sheet_name=today)
            # Upewnij się, że kolumny istnieją
            if "Numer Telefonu" not in existing_df.columns:
                existing_df["Numer Telefonu"] = ""
            if "Odrzucić?" not in existing_df.columns:
                existing_df["Odrzucić?"] = ""
            existing_df["__PhoneNorm__"] = existing_df["Numer Telefonu"].astype(str).apply(_norm_phone)
            existing_decisions = existing_df.set_index("__PhoneNorm__")["Odrzucić?"].to_dict()
        except Exception:
            existing_df = pd.DataFrame(columns=["Branża", "Strona WWW", "Nazwa Firmy", "Adres", "Numer Telefonu", "Odrzucić?"])
            existing_decisions = {}
    else:
        sheet = book.create_sheet(title=today)
        sheet.append(["Branża", "Strona WWW", "Nazwa Firmy", "", "Adres", "", "Numer Telefonu", "Odrzucić?"])
        existing_df = pd.DataFrame(columns=["Branża", "Strona WWW", "Nazwa Firmy", "Adres", "Numer Telefonu", "Odrzucić?"])
        existing_decisions = {}

    # Zbiór WSZYSTKICH numerów z pliku (znormalizowanych)
    all_existing_numbers = set()
    for sheet_name in book.sheetnames:
        try:
            df = pd.read_excel(filename, sheet_name=sheet_name)
            if "Numer Telefonu" in df.columns:
                all_existing_numbers.update(
                    df["Numer Telefonu"].dropna().astype(str).apply(_norm_phone).tolist()
                )
        except Exception:
            continue

    # Ramka z nowymi danymi
    new_df = pd.DataFrame(new_data, columns=["Branża", "Strona WWW", "Nazwa Firmy", "Adres", "Numer Telefonu"])
    new_df["Strona WWW"] = new_df["Strona WWW"].fillna("Brak strony")
    new_df["Numer Telefonu"] = new_df["Numer Telefonu"].astype(str)
    new_df["__PhoneNorm__"] = new_df["Numer Telefonu"].apply(_norm_phone)

    # 1) Usuń duplikaty w bieżącym wsadzie po numerze
    new_df = new_df.drop_duplicates(subset="__PhoneNorm__", keep="first")

    # 2) Usuń rekordy, które już są w pliku (po znormalizowanym numerze)
    new_unique_df = new_df[~new_df["__PhoneNorm__"].isin(all_existing_numbers)].copy()

    # 3) Przenieś wcześniejsze decyzje po kluczu znormalizowanym
    new_unique_df["Odrzucić?"] = new_unique_df["__PhoneNorm__"].map(existing_decisions).fillna("")

    # Walidacja TAK/NIE
    existing_validation = DataValidation(type="list", formula1='"TAK,NIE"', allow_blank=True)
    sheet.add_data_validation(existing_validation)

    # Formatowanie warunkowe – wiersz na czerwono, gdy H=="TAK"
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # Od której linii dopisujemy
    start_row = sheet.max_row + 1

    # Zapis bez kolumny pomocniczej
    for i, (_, r) in enumerate(new_unique_df.iterrows(), start=start_row):
        sheet[f"A{i}"] = r["Branża"]
        if r["Strona WWW"] == "Brak strony":
            sheet[f"B{i}"] = "Brak strony"
        else:
            sheet[f"B{i}"].value = "Kliknij tutaj"
            sheet[f"B{i}"].hyperlink = r["Strona WWW"]
        sheet[f"C{i}"] = r["Nazwa Firmy"]
        sheet[f"E{i}"] = r["Adres"]
        sheet[f"G{i}"] = r["Numer Telefonu"]
        sheet[f"H{i}"] = r["Odrzucić?"]

        # walidacja listy
        existing_validation.add(sheet[f"H{i}"])

        # reguła formatowania dla tego wiersza
        formula = f'$H{i}="TAK"'
        rule = FormulaRule(formula=[formula], fill=red_fill)
        sheet.conditional_formatting.add(f"A{i}:G{i}", rule)

    # Dostosowanie szerokości kolumn
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_len + 2

    # Filtrowanie
    sheet.auto_filter.ref = sheet.dimensions

    book.save(filename)
    logger_util.log_info(f"✅ Dodano {len(new_unique_df)} nowych rekordów do {filename}.")
    # print(f"✅ Dodano {len(new_unique_df)} nowych rekordów do {filename}.")  # wyciszone w GUI
    return len(new_unique_df)